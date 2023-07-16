import pandas as pd
from os import path, makedirs, listdir, rename
from setting import logger, Pattern, Change, Result
from typing import List, Optional
from time import sleep
from tkinter.messagebox import showwarning
from openpyxl.styles import PatternFill, Border, Side
from openpyxl import load_workbook
from openpyxl.styles import Font
from pathlib import Path


def make_dir(dir_names: Path) -> Path:
    try:
        if not path.exists(dir_names):
            makedirs(dir_names)
            logger.info(f"Папка {dir_names} создана")
        return dir_names
    except Exception as ex:
        logger.exception(f"При создании папки: {dir_names}, ошибка: {ex}")


def change(value: str) -> Optional[Change]:
    if value == "Нет":
        return None
    else:
        return Change.YES


def read_pattern(file: Path) -> List[Pattern]:
    pattern = load_workbook(
        file,
        read_only=False,
        keep_vba=False
    )["Вводные данные"]
    task_list: List[Pattern] = []
    for row in pattern.iter_rows(min_row=2):
        task = Pattern(
            url=row[0].value,
            request=row[1].value,
            region=row[2].value,
            type_url=row[3].value,
            commercialization=row[4].value,
            size_top=row[5].value,
            name_task=row[6].value,
        )
        task_list.append(task)
    return task_list


def read_load(file: Path) -> List[Result]:
    pattern_df = pd.read_excel(file, sheet_name='Загруженные отчеты')
    task_list: List[Result] = []
    for row in pattern_df.values.tolist():
        task = Result(
            link=row[0],
            load_tz=change(row[1]),
            recommendations=change(row[2]),
            spam=change(row[3]),
            top_unigrams=change(row[4]),
            top_n_gramm=change(row[5]),
            url=row[6],
            request=row[7],
        )
        task_list.append(task)
    return task_list


def create_file(values: list[tuple], file: Path) -> None:
    try:
        check_busy_file(file)
        workbook = load_workbook(file, read_only=False, keep_vba=False)
        del workbook['Пример входных данных']
        del workbook['Вводные данные']
        workbook['Загруженные отчеты'].sheet_state = 'visible'
        sheet_name = workbook["Загруженные отчеты"]

        i = 2
        for row in values:
            for count in range(8):
                sheet_name.cell(row=i, column=count + 1).value = row[count]
            i += 1
        workbook.save(file)
        logger.info(f"Файл Result.xlsx сформирован и доступен по: {file}")
    except Exception as ex:
        logger.exception(f"При создание файла: {file}, ошибка: {ex}")


def create_result_file(values: list[tuple], file: Path):
    try:
        new_name = path.join(file.parent, f"Result.xlsx")
        check_busy_file(file)
        workbook = load_workbook(file, read_only=False, keep_vba=False)
        del workbook['Загруженные отчеты']
        workbook['Результаты'].sheet_state = 'visible'
        sheetname = workbook['Результаты']

        i = 2
        for row in values:
            for count in range(11):
                sheetname.cell(row=i, column=count + 1).value = row[count]
            i += 1
        workbook.save(new_name)
    except Exception as ex:
        logger.exception(f"При создание файла, ошибка: {ex}")


def wait_file_and_return_path(path_project: Path, name_file: str, how_sek: int = 200) -> Optional[Path]:
    for _ in range(how_sek):
        for files in listdir(path_project):
            if (name_file in files) and (".crdownload" not in files):
                sleep(5)
                return Path(path.join(path_project, files))
        sleep(1)
    logger.info(f"Файл {name_file} не найден, Время ожидания превысило {how_sek} сек.")


def make_dir_project(dir_names: str) -> str:
    if not path.exists(dir_names):
        makedirs(dir_names)
    return dir_names


def get_width_depth_relevance(text: str) -> tuple:
    width = depth = relevance = None
    for row in text.split('\n'):
        if "Охват (ширина облака) = " in row:
            width = row.split(' ')[-1]
        if "ТОП слов (глубина облака) = " in row:
            depth = row.split(' ')[-1]
        if "Общая релевантность =" in row:
            relevance = row.split(' ')[-1]
    return width, depth, relevance


def check_busy_file(file):
    try:
        if path.exists(file):
            rename(file, file)
            return True
    except:
        showwarning(title="Предупреждение", message=f"Файл: {file} занят!\nЗакройте и нажмите ОК")


def add_mean_value(dataframe: pd.DataFrame) -> pd.DataFrame:
    value_for_group = dataframe.iloc[:, 0].unique()
    result_df = pd.DataFrame()
    for value in value_for_group:
        df = dataframe[dataframe.iloc[:, 0] == value]
        mean_value = df.groupby(df.iloc[:, 0]).mean(numeric_only=True).reset_index()
        mean_value.iloc[:, 0] = mean_value.iloc[:, 0] + ' (среднее значение)'
        result_df = pd.concat([result_df, df])
        if df.shape[0] != 1:
            result_df = pd.concat([result_df, mean_value])
    return result_df


def join_sheet_files(files: list[list], sheet_name: str) -> pd.DataFrame:
    dataframe = pd.DataFrame()
    for url, file in files:
        analiz_spam_df = pd.read_excel(file, sheet_name=sheet_name)
        analiz_spam_df['URL'] = url
        dataframe = pd.concat([dataframe, analiz_spam_df])
    return dataframe


def re_spam(url_way_file: list[list]) -> pd.DataFrame:
    columns = ['Слово (самая популярная словоформа)',
               'Повторы у Вас',
               'Минимум повторов (норм.)',
               'Максимум повторов (норм.)',
               'Переспам, %',
               'Переспам * IDF, %',
               'IDF',
               'Количество повторений',
               'URL']
    dataframe = join_sheet_files(url_way_file, 'Переспам')
    count_replay = dataframe.groupby('Слово (самая популярная словоформа)', as_index=False).agg(
        {'URL': 'count'}).rename(columns={'URL': 'Количество повторений'})
    dataframe = dataframe.merge(count_replay).sort_values(by=['Количество повторений',
                                                                                'Слово (самая популярная словоформа)'],
                                                                            ascending=False)
    return add_mean_value(dataframe[columns])


def replay_word(url_way_file: list[list]) -> pd.DataFrame:
    columns = ['Слово (самая популярная словоформа)',
               'Повторы у Вас',
               'Минимум повторов (норм.)',
               'Максимум повторов (норм.)',
               'Количество повторений',
               'URL']
    dataframe = join_sheet_files(url_way_file, 'Повторы слов')
    count_replay = dataframe.groupby('Слово (самая популярная словоформа)', as_index=False).agg(
        {'URL': 'count'}).rename(columns={'URL': 'Количество повторений'})
    dataframe = dataframe.merge(count_replay).sort_values(by=['Количество повторений',
                                                                      'Слово (самая популярная словоформа)'],
                                                                  ascending=False)
    return add_mean_value(dataframe[columns])


def add_common_word(url_way_file: list[list]) -> pd.DataFrame:
    columns = ['Слово (самая популярная словоформа)',
               'Важные словоформы',
               'Все словоформы у конкурентов',
               'Количество повторений',
               'URL']
    dataframe = join_sheet_files(url_way_file, 'Добавить важные слова')
    count_replay = dataframe.groupby('Слово (самая популярная словоформа)', as_index=False).agg(
        {'URL': 'count'}).rename(columns={'URL': 'Количество повторений'})
    dataframe = dataframe.merge(count_replay).sort_values(by=['Количество повторений',
                                                                      'Слово (самая популярная словоформа)'],
                                                                  ascending=False)
    return add_mean_value(dataframe[columns])


def dop_word(url_way_file: list[list]) -> pd.DataFrame:
    columns = ['Дополнительные слова',
               'Количество повторений',
               'URL']

    dataframe = join_sheet_files(url_way_file, 'Доп. слова')
    count_replay = dataframe.groupby('Дополнительные слова',
                                         as_index=False).agg({'URL': 'count'}).rename(
        columns={'URL': 'Количество повторений'})
    dataframe = dataframe.merge(count_replay).sort_values(by=['Количество повторений',
                                                                      'Дополнительные слова'],
                                                                  ascending=False)
    return add_mean_value(dataframe[columns])


def title(url_way_file: list[list]) -> pd.DataFrame:
    columns = ['Можно добавить слова',
               'Количество повторений',
               'URL']
    dataframe = join_sheet_files(url_way_file, 'title')
    count_replay = dataframe.groupby('Можно добавить слова', as_index=False).agg({'URL': 'count'}).rename(
        columns={'URL': 'Количество повторений'})
    dataframe = dataframe.merge(count_replay).sort_values(by=['Количество повторений',
                                                                      'Можно добавить слова'],
                                                                  ascending=False)
    return add_mean_value(dataframe[columns])


def wight_row(path):
    wb = load_workbook(path)
    # размер шрифта документа
    font_size = 10

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # словарь с размерами столбцов
        cols_dict = {}
        list_null_value = [1]

        # проходимся по всем строкам документа
        for row in ws.rows:
            if row[-1].value is None:
                list_null_value.append(row[0].row)
            # теперь по ячейкам каждой строки
            for cell in row:
                # получаем букву текущего столбца
                letter = cell.column_letter
                name_letter = cell.coordinate

                thins = Side(border_style="medium", color="000000")
                ws[name_letter].border = Border(left=thins, right=thins)

                if row[-1].value is None:
                    ws[name_letter].fill = PatternFill('solid', fgColor="F5FFFA")
                    ws[name_letter].border = Border(top=thins, bottom=thins, left=thins, right=thins)
                if cell.row == 1:
                    ws[name_letter].fill = PatternFill('solid', fgColor="9FB6CD")
                    ws[name_letter].border = Border(top=thins, bottom=thins, left=thins, right=thins)

                # если в ячейке записаны данные
                if cell.value:
                    # устанавливаем в ячейке размер шрифта
                    cell.font = Font(name='Calibri', size=font_size)
                    # вычисляем количество символов, записанных в ячейку
                    len_cell = len(str(cell.value))
                    # длинна колонки по умолчанию, если буква
                    # текущего столбца отсутствует в словаре `cols_dict`
                    len_cell_dict = 0
                    # смотрим в словарь c длинами столбцов
                    if letter in cols_dict:
                        # если в словаре есть буква текущего столбца
                        # то извлекаем соответствующую длину
                        len_cell_dict = cols_dict[letter]

                    # если текущая длина данных в ячейке
                    #  больше чем длинна из словаря
                    if len_cell > len_cell_dict:
                        cols_dict[letter] = len_cell
                        new_width_col = len_cell * font_size ** (font_size * 0.009)
                        ws.column_dimensions[cell.column_letter].width = new_width_col

        for count in range(len(list_null_value) - 1):
            min_row = int(list_null_value[count] + 1)
            max_row = int(list_null_value[count + 1] - 1)
            ws.row_dimensions.group(min_row, max_row, hidden=True)

    wb.save(path)